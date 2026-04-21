import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# SHEET 1: RESUMO EXECUTIVO
# ============================================================
ws1 = wb.active
ws1.title = "Resumo Executivo"
ws1.sheet_properties.tabColor = "1F4E79"

title_font = Font(name="Arial", size=16, bold=True, color="1F4E79")
subtitle_font = Font(name="Arial", size=12, bold=True, color="2E75B6")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
alt_fill = PatternFill("solid", fgColor="D6E4F0")
green_fill = PatternFill("solid", fgColor="C6EFCE")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
red_fill = PatternFill("solid", fgColor="FFC7CE")
connected_fill = PatternFill("solid", fgColor="92D050")
body_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", size=10, bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

ws1["A1"] = "EXPLORAÇÃO DE MARKETPLACES - PLUGINS RECOMENDADOS"
ws1["A1"].font = title_font
ws1.merge_cells("A1:F1")

ws1["A3"] = "Status Atual"
ws1["A3"].font = subtitle_font

current = [
    ["Item", "Quantidade", "Status"],
    ["Conectores Cowork ativos", "5", "HubSpot, Google Calendar, Google Drive, Notion, Slack"],
    ["Skills instalados (Cowork)", "33", "Anthropic core skills"],
    ["Skills instalados (Claude Code)", "91+", "dirs com 319 SKILL.md files"],
    ["Plugins disponíveis no registry", "50+", "Marketplace oficial Anthropic"],
    ["Plugins no ecossistema total", "9.000+", "ClaudePluginHub + comunidade"],
]
for i, row in enumerate(current):
    for j, val in enumerate(row):
        cell = ws1.cell(row=4+i, column=j+1, value=val)
        cell.font = header_font if i == 0 else body_font
        cell.fill = header_fill if i == 0 else (alt_fill if i % 2 == 0 else PatternFill())
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 55

ws1["A12"] = "Resumo das Recomendações"
ws1["A12"].font = subtitle_font

summary = [
    ["Prioridade", "Qtd Plugins", "Categoria", "Impacto Esperado"],
    ["ALTA", "8", "Conectores essenciais para C-Level", "Produtividade imediata em vendas, finanças, reuniões"],
    ["MÉDIA", "10", "Ferramentas de desenvolvimento e analytics", "Qualidade de código, dados e insights"],
    ["BAIXA", "7", "Nice-to-have e exploratórios", "Capacidades adicionais quando necessário"],
]
for i, row in enumerate(summary):
    for j, val in enumerate(row):
        cell = ws1.cell(row=13+i, column=j+1, value=val)
        cell.font = header_font if i == 0 else bold_font if j == 0 else body_font
        if i == 0:
            cell.fill = header_fill
        elif i == 1:
            cell.fill = green_fill if j == 0 else PatternFill()
        elif i == 2:
            cell.fill = yellow_fill if j == 0 else PatternFill()
        elif i == 3:
            cell.fill = red_fill if j == 0 else PatternFill()
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws1.column_dimensions["D"].width = 55

# ============================================================
# SHEET 2: CONECTORES COWORK (Registry Anthropic)
# ============================================================
ws2 = wb.create_sheet("Conectores Cowork")
ws2.sheet_properties.tabColor = "2E75B6"

ws2["A1"] = "CONECTORES DISPONÍVEIS NO REGISTRY ANTHROPIC (para Cowork)"
ws2["A1"].font = title_font
ws2.merge_cells("A1:H1")

ws2["A2"] = "Estes conectores podem ser adicionados diretamente no Cowork via Settings > Connectors"
ws2["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
ws2.merge_cells("A2:H2")

headers2 = ["Prioridade", "Conector", "Descrição", "Relevância C-Level", "Categoria", "Tools", "Status", "Ação"]
for j, h in enumerate(headers2):
    cell = ws2.cell(row=4, column=j+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

connectors = [
    # PRIORIDADE ALTA
    ["ALTA", "S&P Global (Kensho)", "Dados financeiros institucionais - receita, capitalização, relações empresariais", "CEO/CFO: Análise de mercado, M&A intelligence, benchmarks", "Finanças", "22 tools", "Disponível", "Conectar"],
    ["ALTA", "FactSet AI-Ready Data", "Dados financeiros premium - FX, bonds, options, interest rates, M&A", "CFO: Valuation, pricing, análise de mercado profissional", "Finanças", "20 tools", "Disponível", "Conectar"],
    ["ALTA", "Apollo.io", "Prospecção B2B - enrich contacts, search people/orgs, job postings", "CMO/CEO: Pipeline de vendas, lead generation, outbound", "Vendas", "13 tools", "Disponível", "Conectar"],
    ["ALTA", "Circleback", "Context de reuniões - transcripts, meetings, profiles, emails", "CEO: Revisão de meetings, follow-ups, action items", "Reuniões", "9 tools", "Disponível", "Conectar"],
    ["ALTA", "Fellow.ai", "Meeting insights - summaries, transcripts, action items, participants", "CEO/CPO: Gestão de reuniões e decisões", "Reuniões", "5 tools", "Disponível", "Conectar"],
    ["ALTA", "Atlassian (Jira+Confluence)", "Gestão de projetos e documentação técnica", "CPO/CEO: Tracking de projetos, roadmaps, docs", "Projetos", "31 tools", "Disponível", "Conectar"],
    ["ALTA", "Linear", "Issue tracking moderno - gestão de tickets, projetos, ciclos", "CPO: Sprint management, product development", "Projetos", "22 tools", "Disponível", "Conectar"],
    ["ALTA", "Asana", "Coordenação de tasks, projetos e goals", "CEO/CPO: Portfolio management, team alignment", "Projetos", "15 tools", "Disponível", "Conectar"],
    # PRIORIDADE MÉDIA
    ["MÉDIA", "Klaviyo", "Email marketing - campaigns, flows, segments, metrics", "CMO: Email automation, customer engagement", "Marketing", "27 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "ActiveCampaign", "Marketing automation - contacts, campaigns, automations, deals", "CMO: Nurturing, automation, CRM leve", "Marketing", "32 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "Amplitude", "Product analytics - charts, experiments, datasets, metrics", "CPO/CMO: Product usage, experimentation, funnel analysis", "Analytics", "14 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "PostHog", "Product analytics + feature flags + session replay", "CPO: Product decisions baseadas em dados", "Analytics", "57 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "Ahrefs", "SEO analytics - keywords, backlinks, content gaps, site audit", "CMO: SEO strategy, competitor analysis, content", "SEO", "61 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "Brex", "Gestão financeira - expenses, cards, limits, reimbursements", "CFO: Controle de gastos, compliance, aprovações", "Finanças", "10 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "Windsor.ai", "325+ data sources de marketing, analytics e CRM unificados", "CMO: Marketing attribution, cross-channel analytics", "Marketing", "5 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "Vibe Prospecting", "Enrichment de empresas e prospects com dados Explorium", "CMO/CEO: Lead scoring, account intelligence", "Vendas", "12 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "ZoomInfo", "Intelligence de contacts e accounts para GTM", "CMO: Enrichment, prospecting, competitive intel", "Vendas", "5 tools", "Disponível", "Avaliar"],
    ["MÉDIA", "Calendly", "Scheduling - event types, availability, bookings", "CEO: Agenda management, automated scheduling", "Produtividade", "34 tools", "Disponível", "Avaliar"],
    # PRIORIDADE BAIXA
    ["BAIXA", "Granola", "AI notepad para meetings", "CEO: Meeting notes automáticas", "Reuniões", "4 tools", "Disponível", "Opcional"],
    ["BAIXA", "Clockwise", "Time management e scheduling inteligente", "CEO: Otimização de tempo, focus time", "Produtividade", "15 tools", "Disponível", "Opcional"],
    ["BAIXA", "Visier", "People analytics - métricas de produtividade e impacto", "CEO: HR analytics, workforce planning", "HR/Analytics", "8 tools", "Disponível", "Opcional"],
    ["BAIXA", "Pigment", "Business planning e análise de dados", "CFO: Business planning, forecasting", "Finanças", "3 tools", "Disponível", "Opcional"],
    ["BAIXA", "Day AI", "CRMx - AI-native CRM", "CEO: Alternative CRM com AI", "CRM", "25 tools", "Disponível", "Opcional"],
    ["BAIXA", "Omni Analytics", "Natural language data queries", "CFO: Self-service analytics", "Analytics", "3 tools", "Disponível", "Opcional"],
    ["BAIXA", "Udemy Business", "Learning & development - courses, paths, labs", "CEO: Team upskilling, L&D", "Educação", "5 tools", "Disponível", "Opcional"],
]

for i, row in enumerate(connectors):
    for j, val in enumerate(row):
        cell = ws2.cell(row=5+i, column=j+1, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if j == 0:
            cell.font = bold_font
            if val == "ALTA":
                cell.fill = green_fill
            elif val == "MÉDIA":
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill

widths2 = [12, 24, 55, 50, 14, 12, 12, 12]
for j, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(j+1)].width = w

# ============================================================
# SHEET 3: PLUGINS CLAUDE CODE (Terminal)
# ============================================================
ws3 = wb.create_sheet("Plugins Claude Code")
ws3.sheet_properties.tabColor = "548235"

ws3["A1"] = "PLUGINS PARA CLAUDE CODE (Terminal - /plugin install)"
ws3["A1"].font = title_font
ws3.merge_cells("A1:H1")

ws3["A2"] = "Estes plugins são instalados via terminal do Claude Code com /plugin install"
ws3["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
ws3.merge_cells("A2:H2")

headers3 = ["Prioridade", "Plugin", "Descrição", "Installs", "Comando de Instalação", "Relevância", "Categoria", "Status"]
for j, h in enumerate(headers3):
    cell = ws3.cell(row=4, column=j+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

plugins_cc = [
    ["ALTA", "Context7", "Documentação real-time de bibliotecas - elimina hallucinations", "71.800+", "/plugin install context7@claude-plugins-official", "Dev: Código correto com APIs atualizadas", "Dev", "Instalar"],
    ["ALTA", "Frontend Design", "Melhora UI/UX de interfaces geradas por AI", "96.400+", "/plugin install frontend-design@claude-plugins-official", "CPO/CMO: Protótipos profissionais", "Design", "Instalar"],
    ["ALTA", "Code Review", "Review automatizado com múltiplos agentes especializados", "50.000+", "/plugin install code-review@claude-plugins-official", "CTO: Qualidade e segurança do código", "Dev", "Instalar"],
    ["ALTA", "Ralph Loop", "Autonomous coding sessions com commits automáticos", "57.000+", "/plugin install ralph-loop@claude-plugins-official", "Dev: Automação de tarefas repetitivas", "Dev", "Já instalado"],
    ["ALTA", "Security Guidance", "Scanner de vulnerabilidades antes de commit", "25.500+", "/plugin install security-guidance@claude-plugins-official", "CTO/CEO: Segurança do código", "Security", "Instalar"],
    ["MÉDIA", "Playwright", "Browser automation e testing via natural language", "28.100+", "/plugin install playwright@claude-plugins-official", "CPO: Testing E2E automatizado", "Testing", "Instalar"],
    ["MÉDIA", "Figma MCP", "Design-to-code - lê Figma e gera componentes", "18.100+", "/plugin install figma@claude-plugins-official", "CPO/Design: Figma → código funcional", "Design", "Instalar"],
    ["MÉDIA", "Linear Plugin", "Issue tracking integrado ao desenvolvimento", "9.500+", "/plugin install linear@claude-plugins-official", "CPO: Dev workflow integrado", "Projetos", "Instalar"],
    ["MÉDIA", "Firecrawl", "Web scraping - qualquer site → dados limpos para AI", "N/A", "/plugin install firecrawl@claude-plugins-official", "Todos: Extração de dados web", "Data", "Já instalado"],
    ["MÉDIA", "Superpowers", "Lifecycle planning + skills: brainstorm, TDD, debug, review", "N/A", "Via GitHub", "Dev: Workflow estruturado de desenvolvimento", "Dev", "Avaliar"],
    ["MÉDIA", "Chrome DevTools MCP", "Debug live pages - network, console, performance", "20.000+", "/plugin install chrome-devtools-mcp@chrome-devtools-plugins", "Dev: Debug em browser real", "Dev", "Instalar"],
    ["BAIXA", "Shipyard", "IaC validation + security audit (Terraform, Docker, K8s)", "N/A", "Via GitHub", "CTO: Infraestrutura segura", "DevOps", "Avaliar"],
    ["BAIXA", "TypeScript LSP", "Type checking real via LSP integrado", "N/A", "Via GitHub", "Dev: Erros de tipo em tempo real", "Dev", "Avaliar"],
    ["BAIXA", "Agent Peer Review", "Multi-model review (Claude + Codex)", "N/A", "Via GitHub", "Dev: Review cruzado AI", "Dev", "Avaliar"],
    ["BAIXA", "Plannotator", "Planning mode aprimorado com anotações", "N/A", "Via GitHub", "Dev: Planos mais claros", "Dev", "Avaliar"],
]

for i, row in enumerate(plugins_cc):
    for j, val in enumerate(row):
        cell = ws3.cell(row=5+i, column=j+1, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if j == 0:
            cell.font = bold_font
            if val == "ALTA":
                cell.fill = green_fill
            elif val == "MÉDIA":
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill
        if j == 7:
            if val == "Já instalado":
                cell.fill = connected_fill
                cell.font = bold_font

widths3 = [12, 22, 50, 12, 52, 42, 12, 14]
for j, w in enumerate(widths3):
    ws3.column_dimensions[get_column_letter(j+1)].width = w

# ============================================================
# SHEET 4: PLANO DE AÇÃO
# ============================================================
ws4 = wb.create_sheet("Plano de Ação")
ws4.sheet_properties.tabColor = "BF8F00"

ws4["A1"] = "PLANO DE AÇÃO - INSTALAÇÃO POR FASES"
ws4["A1"].font = title_font
ws4.merge_cells("A1:F1")

# FASE 1
ws4["A3"] = "FASE 1: Quick Wins (Esta semana)"
ws4["A3"].font = subtitle_font
ws4["A3"].fill = green_fill

phase1_headers = ["#", "Ação", "Onde", "Tempo", "Impacto", "Como"]
for j, h in enumerate(phase1_headers):
    cell = ws4.cell(row=4, column=j+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

phase1 = [
    ["1", "Conectar S&P Global", "Cowork Settings", "2 min", "CFO: Dados financeiros premium", "Settings > Connectors > S&P Global > Connect"],
    ["2", "Conectar FactSet", "Cowork Settings", "2 min", "CFO: Valuation e M&A data", "Settings > Connectors > FactSet > Connect"],
    ["3", "Conectar Apollo.io", "Cowork Settings", "2 min", "CMO: Lead generation", "Settings > Connectors > Apollo > Connect"],
    ["4", "Conectar Atlassian", "Cowork Settings", "2 min", "CPO: Jira + Confluence", "Settings > Connectors > Atlassian > Connect"],
    ["5", "Conectar Circleback", "Cowork Settings", "2 min", "CEO: Meeting intelligence", "Settings > Connectors > Circleback > Connect"],
    ["6", "Conectar Fellow.ai", "Cowork Settings", "2 min", "CEO: Meeting insights", "Settings > Connectors > Fellow > Connect"],
    ["7", "Instalar Context7", "Claude Code terminal", "1 min", "Dev: Docs atualizadas", "/plugin install context7@claude-plugins-official"],
    ["8", "Instalar Frontend Design", "Claude Code terminal", "1 min", "Design: UI profissional", "/plugin install frontend-design@claude-plugins-official"],
    ["9", "Instalar Code Review", "Claude Code terminal", "1 min", "Dev: Quality assurance", "/plugin install code-review@claude-plugins-official"],
    ["10", "Instalar Security Guidance", "Claude Code terminal", "1 min", "Security: Scan pré-commit", "/plugin install security-guidance@claude-plugins-official"],
]

for i, row in enumerate(phase1):
    for j, val in enumerate(row):
        cell = ws4.cell(row=5+i, column=j+1, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# FASE 2
ws4["A17"] = "FASE 2: Integrações de Marketing e Analytics (Próxima semana)"
ws4["A17"].font = subtitle_font
ws4["A17"].fill = yellow_fill

for j, h in enumerate(phase1_headers):
    cell = ws4.cell(row=18, column=j+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

phase2 = [
    ["1", "Conectar Klaviyo", "Cowork Settings", "5 min", "CMO: Email marketing", "Settings > Connectors > Klaviyo > Connect"],
    ["2", "Conectar Ahrefs", "Cowork Settings", "5 min", "CMO: SEO intelligence", "Settings > Connectors > Ahrefs > Connect"],
    ["3", "Conectar PostHog ou Amplitude", "Cowork Settings", "5 min", "CPO: Product analytics", "Escolher um dos dois baseado no stack atual"],
    ["4", "Conectar Brex", "Cowork Settings", "5 min", "CFO: Expense management", "Settings > Connectors > Brex > Connect"],
    ["5", "Instalar Playwright", "Claude Code terminal", "2 min", "Dev: Browser testing", "/plugin install playwright@claude-plugins-official"],
    ["6", "Instalar Figma MCP", "Claude Code terminal", "2 min", "Design: Figma → código", "/plugin install figma@claude-plugins-official"],
    ["7", "Conectar Calendly", "Cowork Settings", "5 min", "CEO: Scheduling automático", "Settings > Connectors > Calendly > Connect"],
    ["8", "Conectar Windsor.ai", "Cowork Settings", "5 min", "CMO: Marketing attribution", "Settings > Connectors > Windsor > Connect"],
]

for i, row in enumerate(phase2):
    for j, val in enumerate(row):
        cell = ws4.cell(row=19+i, column=j+1, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# FASE 3
ws4["A29"] = "FASE 3: Otimização e Exploratórios (Quando necessário)"
ws4["A29"].font = subtitle_font

for j, h in enumerate(phase1_headers):
    cell = ws4.cell(row=30, column=j+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

phase3 = [
    ["1", "Conectar Visier", "Cowork Settings", "5 min", "CEO: People analytics", "Quando precisar de HR analytics"],
    ["2", "Conectar Granola", "Cowork Settings", "2 min", "CEO: AI meeting notes", "Se usar Granola como notetaker"],
    ["3", "Instalar Chrome DevTools MCP", "Claude Code terminal", "2 min", "Dev: Live debugging", "Quando debuggar apps web"],
    ["4", "Avaliar Superpowers plugin", "Claude Code terminal", "10 min", "Dev: Workflow estruturado", "Testar em projeto piloto"],
    ["5", "Criar Plugin Nuvini", "Claude Code terminal", "2-4h", "Todos: Custom workflows", "Usar skill create-cowork-plugin"],
    ["6", "Explorar marketplaces comunitários", "Web", "1h", "Todos: Discovery contínuo", "ClaudePluginHub, buildwithclaude.com"],
]

for i, row in enumerate(phase3):
    for j, val in enumerate(row):
        cell = ws4.cell(row=31+i, column=j+1, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 32
ws4.column_dimensions["C"].width = 22
ws4.column_dimensions["D"].width = 10
ws4.column_dimensions["E"].width = 32
ws4.column_dimensions["F"].width = 52

# ============================================================
# SHEET 5: MARKETPLACES DISPONÍVEIS
# ============================================================
ws5 = wb.create_sheet("Marketplaces")
ws5.sheet_properties.tabColor = "7030A0"

ws5["A1"] = "ECOSSISTEMA DE MARKETPLACES CLAUDE CODE"
ws5["A1"].font = title_font
ws5.merge_cells("A1:E1")

headers5 = ["Marketplace", "Plugins", "Tipo", "URL", "Notas"]
for j, h in enumerate(headers5):
    cell = ws5.cell(row=3, column=j+1, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

marketplaces = [
    ["Anthropic Official", "50+", "Oficial", "claude-plugins-official (built-in)", "Curado pela Anthropic, alta qualidade"],
    ["Cowork Connectors Registry", "50+", "Oficial", "Settings > Connectors (built-in)", "MCP servers gerenciados"],
    ["ClaudePluginHub", "9.000+", "Comunidade", "claudepluginhub.com", "Maior agregador da comunidade"],
    ["Build with Claude", "488+", "Comunidade", "buildwithclaude.com", "Marketplace com reviews"],
    ["SkillsMP", "400.000+", "Comunidade", "skillsmp.com", "Skills e prompts genéricos"],
    ["Awesome Claude Plugins (Chat2AnyLLM)", "Curado", "GitHub", "github.com/Chat2AnyLLM/awesome-claude-plugins", "Lista curada de marketplaces"],
    ["Awesome Claude Plugins (quemsah)", "Curado", "GitHub", "github.com/quemsah/awesome-claude-plugins", "Métricas de adoção"],
    ["Anthropic Official GitHub", "Curado", "GitHub", "github.com/anthropics/claude-plugins-official", "Plugins oficiais de alta qualidade"],
    ["Composio", "500+", "Plataforma", "composio.dev", "Integração com 500+ SaaS apps"],
    ["Repomix", "Guia", "Docs", "repomix.com/guide/claude-code-plugins", "Documentação e guias"],
    ["Claude Marketplaces", "Agregador", "Web", "claudemarketplaces.com", "Diretório de marketplaces"],
]

for i, row in enumerate(marketplaces):
    for j, val in enumerate(row):
        cell = ws5.cell(row=4+i, column=j+1, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if i % 2 == 0:
            cell.fill = alt_fill

ws5.column_dimensions["A"].width = 36
ws5.column_dimensions["B"].width = 14
ws5.column_dimensions["C"].width = 14
ws5.column_dimensions["D"].width = 48
ws5.column_dimensions["E"].width = 40

# Save
output = "/sessions/trusting-compassionate-maxwell/mnt/Claude/Exploracao_Marketplaces_Plugins.xlsx"
wb.save(output)
print(f"Saved to {output}")
